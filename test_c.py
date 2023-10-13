import re
import os
import openpyxl
import xlrd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import selenium
import time
 
def readfilename(path):
    papername = re.compile(r"(\d+).*?\s(\w.*)")
    Papers = []
    for filename in os.listdir(path):
        paper = papername.search(filename[:-4])
        if paper:
            Papers.append([int(paper.group(1)), paper.group(2)])
    # print(Papers)
    Papers.sort()
    for paper in Papers:
        print(paper)
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = '1204'
    sheet['A1'] = "序号"
    sheet['B1'] = "论文名称"
    sheet['C1'] = '出版时间'
    sheet['D1'] = '出版社'
    for i in range(len(Papers)):
        sheet.cell(row=i + 2, column=1).value, sheet.cell(row=i + 2, column=2).value = Papers[i]
    wb.save('论文信息统计.xlsx')
 
    return Papers
 
 
def notrobort(wd, File):
    """出现人机认证时，激活该部分，先手动通过图片认证"""
 
    element = wd.find_element(By.ID, 'gs_hdr_tsi')
    element.send_keys(File[0][1] + '\n')
    time.sleep(15)  # 用于图片认证的时间 10s
    linkElem = wd.find_element(By.LINK_TEXT, '引用')
    linkElem.click()
    linkElem = wd.find_element(By.LINK_TEXT, 'BibTeX')
    linkElem.click()
    element = wd.find_element(By.TAG_NAME, 'pre')
    print(element.text)
    time.sleep(2)
    wd.refresh()
    wd.back()
    wd.back()
    wd.back()
    wd.refresh()
    wd.back()
 
 
def bib(wd, index, l, File):
    with open('bib.txt', 'a', encoding='utf-8') as f:
        while index < l:
            try:
                element = wd.find_element(By.ID, 'gs_hdr_tsi')
                element.send_keys(File[index] + '\n')
                linkElem = wd.find_element(By.LINK_TEXT, '引用')
                linkElem.click()
                linkElem = wd.find_element(By.LINK_TEXT, 'BibTeX')
                linkElem.click()
                element = wd.find_element(By.TAG_NAME, 'pre')
                print(element.text)
                f.write('\n' + element.text)
 
                time.sleep(2)
                wd.refresh()
                wd.back()
                wd.back()
                wd.back()
                index += 1
            except selenium.common.exceptions.NoSuchElementException:
                wd.quit()
                key = input("请按 Y 进行人机验证\n")
                if key == 'Y':
                    print('将进行人机验证\n')
                    option = webdriver.ChromeOptions()
                    option.add_experimental_option("detach", True)
                    wd = webdriver.Chrome(
                        service=Service(r'C:\Program Files\Google\Chrome\Application\chromedriver.exe'),
                        options=option)
                    wd.implicitly_wait(5)
                    wd.get('https://scholar.google.com/')
                    wd.maximize_window()
                    notrobort(wd, File)
        wd.quit()
 
 
def bibdownload():
    # File = readfilename(path)
    readfile = xlrd.open_workbook(r"C:/Users/vincent/Desktop/1.xlsx")
    print(readfile)
    names = readfile.sheet_names()
    print(names)
    obj_sheet = readfile.sheet_by_name("Sheet1")
    print(obj_sheet)
    # col = obj_sheet.ncols
    # 获取sheet行数
    row = obj_sheet.nrows
    # 获取sheet列数
    col = obj_sheet.ncols
    print("row:", row)
    print("col:", col)
    
    l = len(obj_sheet.col_values(0))
    index = 58
    option = webdriver.ChromeOptions()
    option.add_experimental_option("detach", True)
    wd = webdriver.Chrome(executable_path='C:\Program Files\Google\Chrome\Application\chromedriver.exe',
                          options=option)
    wd.implicitly_wait(5)
    wd.get('https://scholar.google.com/')
    wd.maximize_window()
    # print(obj_sheet.col_values(0))
    bib(wd, index, 76, obj_sheet.col_values(0))
 
 
    wd = webdriver.Chrome(service=Service(r'C:\Program Files\Google\Chrome\Application\chromedriver.exe'),
                          options=option)
    wd.implicitly_wait(5)
    wd.get('https://www.apple.com.cn/')
    wd.maximize_window()
    time.sleep(12)
    wd.quit()
 
 
def openreadtxt(file_name):
    data = []
    file = open(file_name, 'r')  # 打开文件
    file_data = file.readlines()  # 读取所有行
    for row in file_data:
        data.append(row[:-1])  # 将每行数据插入data中
    return data
 
def paperinfo():
    data = openreadtxt('bib.txt')
    time = re.compile(r"year={(\d\d\d\d)}")
    publisher = re.compile(r'(journal|booktitle)={(.*?)}')
    T = []
    Pub = []
    for row in data:
        m = time.search(row)
        h = publisher.search(row)
        if m:
            print(m.group(1))
            T.append(int(m.group(1)))
        if h:
            print(h.group(2))
            Pub.append(h.group(2))
    print(len(T))
    print(len(Pub))
    wb = openpyxl.load_workbook('论文信息统计.xlsx')
    sheet = wb['1204']
    for i in range(len(Pub)):
        sheet.cell(row=i + 2, column=3).value = T[i]
        sheet.cell(row=i + 2, column=4).value = Pub[i]
    wb.save('论文信息统计.xlsx')
 
 
def paperrush():
    bibdownload()
    paperinfo()
 
 
if __name__ == "__main__":
    path = r'D:\GridOcean\paper'
    paperrush()